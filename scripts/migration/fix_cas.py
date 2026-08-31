from __future__ import annotations

import re
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook


def clean_text(s: str) -> str:
    """清理单元格里的换行、引号、全角空格、普通空格。"""
    s = str(s)
    s = s.replace("\n", "").replace("\r", "")
    s = s.replace('"', "").replace("“", "").replace("”", "")
    s = s.replace("'", "")
    s = s.replace("\u3000", " ").replace("\xa0", " ")
    return s.strip()


def extract_ymd(value: Any) -> tuple[int, int, int] | None:
    """
    从 Excel 单元格值中提取 year, month, day。
    支持：
    - datetime/date
    - '1979/1/6'
    - '1979-1-6'
    - '1979.1.6'
    """
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.year, value.month, value.day

    if isinstance(value, date):
        return value.year, value.month, value.day

    s = clean_text(value)
    if not s:
        return None

    # 常见分隔符统一
    s = s.replace(".", "/").replace("-", "/")
    m = re.fullmatch(r"(\d{1,4})/(\d{1,2})/(\d{1,2})", s)
    if not m:
        return None

    y = int(m.group(1))
    mo = int(m.group(2))
    d = int(m.group(3))

    if not (1 <= mo <= 12 and 1 <= d <= 31):
        return None

    return y, mo, d


def cas_checksum_ok(cas_number: str) -> bool:
    """
    校验 CAS。
    规则：去掉连字符后，最后一位为校验位。
    前面的数字从右往左乘 1,2,3... 求和，对 10 取余。
    """
    s = cas_number.strip()
    if not re.fullmatch(r"\d{2,7}-\d{2}-\d", s):
        return False

    digits = s.replace("-", "")
    body = digits[:-1]
    check_digit = int(digits[-1])

    total = 0
    for i, ch in enumerate(reversed(body), start=1):
        total += int(ch) * i

    return total % 10 == check_digit


def build_cas_4(y: int, m: int, d: int) -> str:
    return f"{y}-{m:02d}-{d}"


def build_cas_2(y: int, m: int, d: int) -> str:
    return f"{y % 100}-{m:02d}-{d}"


def recover_cas(value: Any) -> dict[str, str]:
    """
    恢复逻辑：
    - 非 19xx：直接输出四位候选
    - 19xx：同时校验四位候选和两位候选
        * 只有一个合法 -> 输出该合法项
        * 两个都合法 / 都不合法 -> 待确认
    """
    ymd = extract_ymd(value)
    if ymd is None:
        return {
            "恢复CAS": "",
            "四位候选": "",
            "两位候选": "",
            "判断": "无法解析"
        }

    y, m, d = ymd
    cas4 = build_cas_4(y, m, d)
    cas2 = build_cas_2(y, m, d)

    # 只对 19xx 做歧义判断
    if 1900 <= y <= 1999:
        v4 = cas_checksum_ok(cas4)
        v2 = cas_checksum_ok(cas2)

        if v4 and not v2:
            result = cas4
            reason = "19xx：四位合法，两位不合法"
        elif v2 and not v4:
            result = cas2
            reason = "19xx：两位合法，四位不合法"
        elif v2 and v4:
            result = "待确认"
            reason = "19xx：两位和四位都合法"
        else:
            result = "待确认"
            reason = "19xx：两位和四位都不合法"
    else:
        result = cas4
        reason = "非19xx：直接按四位恢复"

    return {
        "恢复CAS": result,
        "四位候选": cas4,
        "两位候选": cas2,
        "判断": reason
    }


def process_excel(
    input_file: str,
    output_file: str,
    cas_col: int = 1,
    name_col: int = 2,
    header_row: int = 1,
) -> None:
    """
    处理 Excel：
    - cas_col: 被错误转成日期的 CAS 列，默认第1列(A)
    - name_col: 名称列，默认第2列(B)
    - header_row: 表头行，默认第1行
    """
    wb = load_workbook(input_file)
    ws = wb.active

    # 新增输出列
    out_col_result = ws.max_column + 1
    out_col_cas4 = ws.max_column + 2
    out_col_cas2 = ws.max_column + 3
    out_col_reason = ws.max_column + 4

    ws.cell(header_row, out_col_result, "恢复CAS")
    ws.cell(header_row, out_col_cas4, "四位候选")
    ws.cell(header_row, out_col_cas2, "两位候选")
    ws.cell(header_row, out_col_reason, "判断")

    for row in range(header_row + 1, ws.max_row + 1):
        cas_value = ws.cell(row, cas_col).value
        result = recover_cas(cas_value)

        ws.cell(row, out_col_result, result["恢复CAS"])
        ws.cell(row, out_col_cas4, result["四位候选"])
        ws.cell(row, out_col_cas2, result["两位候选"])
        ws.cell(row, out_col_reason, result["判断"])

    wb.save(output_file)


if __name__ == "__main__":
    input_path = "input.xlsx"
    output_path = "output_fixed.xlsx"

    # 默认：
    # A列 = 错误CAS
    # B列 = 名称
    # 第1行 = 表头
    process_excel(
        input_file=input_path,
        output_file=output_path,
        cas_col=1,
        name_col=2,
        header_row=1,
    )

    print(f"已完成，输出文件：{output_path}")