"""XLSX export service for unified export functionality."""
from __future__ import annotations

import io
from dataclasses import dataclass
from enum import Enum
from typing import Any

from fastapi import HTTPException, status
from fastapi.responses import StreamingResponse

from app.core.time_utils import get_display_now, to_display_time
from app.services.spec_utils import format_specification

_DANGEROUS_SPREADSHEET_PREFIXES = ("=", "+", "-", "@")

try:
    from openpyxl import Workbook
    from openpyxl.styles.numbers import FORMAT_TEXT
except ImportError as exc:  # pragma: no cover - import-time guard for deployment envs
    raise HTTPException(
        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        detail="XLSX export requires the 'openpyxl' package.",
    ) from exc


@dataclass
class ExportSheet:
    title: str
    headers: list[str]
    rows: list[list[Any]]
    text_columns: set[int]


def _format_display_datetime(value: Any) -> str:
    display_time = to_display_time(value)
    if display_time is None:
        return ""
    return display_time.strftime("%Y-%m-%d %H:%M:%S")


def _get_field(item: Any, field: str, default: Any = None) -> Any:
    if isinstance(item, dict):
        return item.get(field, default)
    return getattr(item, field, default)


def _escape_spreadsheet_formula(value: Any) -> Any:
    """Neutralize spreadsheet formulas in exported string cells."""
    if not isinstance(value, str):
        return value

    stripped = value.lstrip()
    if stripped and stripped[0] in _DANGEROUS_SPREADSHEET_PREFIXES:
        return f"'{value}"

    return value


def _unwrap_enum_value(value: Any) -> Any:
    if isinstance(value, Enum):
        return value.value
    return value


def _export_to_xlsx(
    sheets: list[ExportSheet],
    filename_prefix: str,
) -> StreamingResponse:
    """Generate XLSX export response with per-column text formatting."""
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for sheet in sheets:
        ws = wb.create_sheet(title=sheet.title)

        for col_idx, header in enumerate(sheet.headers, 1):
            header_cell = ws.cell(row=1, column=col_idx, value=header)
            if col_idx in sheet.text_columns:
                header_cell.number_format = FORMAT_TEXT

        for row_idx, row in enumerate(sheet.rows, 2):
            for col_idx, value in enumerate(row, 1):
                safe_value = _escape_spreadsheet_formula(value)
                if col_idx in sheet.text_columns:
                    cell = ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value="" if safe_value is None else str(safe_value),
                    )
                    cell.number_format = FORMAT_TEXT
                else:
                    ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value="" if safe_value is None else safe_value,
                    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{filename_prefix}_{get_display_now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def export_inventory_xlsx(
    items: list[Any],
) -> StreamingResponse:
    """Export inventory workbook with regular inventory only."""
    regular_headers = [
        "CAS号",
        "名称",
        "英文名",
        "别名",
        "分类",
        "品牌",
        "纯度",
        "位置",
        "初始数量",
        "剩余数量",
        "单位",
        "状态",
        "是否危险品",
        "入库时间",
        "备注",
    ]

    def regular_row_converter(item: Any) -> list[Any]:
        return [
            _get_field(item, "cas_number"),
            _get_field(item, "name"),
            _get_field(item, "english_name") or "",
            _get_field(item, "alias") or "",
            _get_field(item, "category") or "",
            _get_field(item, "brand") or "",
            _get_field(item, "purity") or "",
            _get_field(item, "storage_location") or "",
            _get_field(item, "initial_quantity"),
            _get_field(item, "remaining_quantity"),
            _get_field(item, "unit"),
            _get_field(item, "status").value
            if hasattr(_get_field(item, "status"), "value")
            else _get_field(item, "status"),
            "是" if _get_field(item, "is_hazardous") else "否",
            _format_display_datetime(_get_field(item, "created_at")),
            _get_field(item, "notes") or "",
        ]

    regular_rows = [regular_row_converter(item) for item in items]
    sheets = [
        ExportSheet(
            title="库存",
            headers=regular_headers,
            rows=regular_rows,
            text_columns=set(range(1, len(regular_headers) + 1)),
        )
    ]

    return _export_to_xlsx(
        sheets=sheets,
        filename_prefix="inventory_export",
    )


def export_common_shelf_xlsx(
    groups: list[Any],
) -> StreamingResponse:
    """Export grouped common shelf data."""
    headers = [
        "CAS号",
        "名称",
        "英文名",
        "分类",
        "品牌",
        "纯度",
        "规格",
        "剩余瓶数",
        "位置数",
        "最新入库名称",
        "备注",
        "创建时间",
        "更新时间",
    ]

    def row_converter(group: Any) -> list[Any]:
        group_data = _get_field(group, "group", {}) or {}
        display_data = _get_field(group, "display", {}) or {}
        created_at = _get_field(group, "created_at")
        updated_at = _get_field(group, "updated_at")
        return [
            _get_field(group_data, "cas_number"),
            _get_field(display_data, "name"),
            _get_field(display_data, "english_name") or "",
            _unwrap_enum_value(_get_field(display_data, "category")) or "",
            _get_field(group_data, "brand") or "",
            _get_field(display_data, "purity") or "",
            _get_field(group_data, "specification_text") or "",
            _get_field(group, "bottle_count") or 0,
            _get_field(group, "location_count") or 0,
            _get_field(group, "latest_name_snapshot") or "",
            _get_field(display_data, "notes") or "",
            _format_display_datetime(created_at),
            _format_display_datetime(updated_at),
        ]

    rows = [row_converter(group) for group in groups]
    text_columns = {1, 2, 3, 4, 5, 6, 7, 10, 11, 12, 13}
    return _export_to_xlsx(
        sheets=[
            ExportSheet(
                title="常用货架",
                headers=headers,
                rows=rows,
                text_columns=text_columns,
            )
        ],
        filename_prefix="common_shelf_export",
    )


def export_reagent_orders_xlsx(
    items: list[Any],
    users_map: dict[int, str] | None = None,
) -> StreamingResponse:
    """Export reagent orders."""
    headers = [
        "CAS号",
        "名称",
        "英文名",
        "别名",
        "分类",
        "品牌",
        "纯度",
        "规格",
        "数量",
        "单价",
        "申购原因",
        "状态",
        "是否危险品",
        "订购人",
        "申购时间",
        "备注",
    ]

    def row_converter(item: Any, user_map: dict[int, str]) -> list[Any]:
        status_value = _get_field(item, "status")
        order_reason_value = _get_field(item, "order_reason")
        initial_quantity = _get_field(item, "initial_quantity")
        unit = _get_field(item, "unit")
        specification = _get_field(item, "specification")
        if not specification:
            specification = format_specification(initial_quantity, unit)

        return [
            _get_field(item, "cas_number"),
            _get_field(item, "name"),
            _get_field(item, "english_name") or "",
            _get_field(item, "alias") or "",
            _get_field(item, "category") or "",
            _get_field(item, "brand") or "",
            _get_field(item, "purity") or "",
            specification or "",
            _get_field(item, "quantity"),
            _get_field(item, "price") or "",
            order_reason_value.value if hasattr(order_reason_value, "value") else order_reason_value,
            status_value.value if hasattr(status_value, "value") else status_value,
            "是" if _get_field(item, "is_hazardous") else "否",
            user_map.get(_get_field(item, "applicant_id"), "") if _get_field(item, "applicant_id") else "",
            _format_display_datetime(_get_field(item, "created_at")),
            _get_field(item, "notes") or "",
        ]

    resolved_users_map = users_map or {}
    rows = [row_converter(item, resolved_users_map) for item in items]
    # 单价列保持数值格式，其余列全部强制文本格式
    text_columns = {idx for idx in range(1, len(headers) + 1) if idx != 10}
    return _export_to_xlsx(
        sheets=[
            ExportSheet(
                title="试剂订单",
                headers=headers,
                rows=rows,
                text_columns=text_columns,
            )
        ],
        filename_prefix="reagent_orders_export",
    )


def export_consumable_orders_xlsx(
    items: list[Any],
    users_map: dict[int, str] | None = None,
) -> StreamingResponse:
    """Export consumable orders."""
    headers = [
        "名称",
        "英文名",
        "规格",
        "数量",
        "单价",
        "状态",
        "订购人",
        "申购时间",
        "备注",
    ]

    def row_converter(item: Any, user_map: dict[int, str]) -> list[Any]:
        status_value = _get_field(item, "status")
        spec = _get_field(item, "specification", "") or ""
        return [
            _get_field(item, "name"),
            _get_field(item, "english_name") or "",
            spec,
            _get_field(item, "quantity"),
            _get_field(item, "price") or "",
            status_value.value if hasattr(status_value, "value") else status_value,
            user_map.get(_get_field(item, "applicant_id"), "") if _get_field(item, "applicant_id") else "",
            _format_display_datetime(_get_field(item, "created_at")),
            _get_field(item, "notes") or "",
        ]

    resolved_users_map = users_map or {}
    rows = [row_converter(item, resolved_users_map) for item in items]
    # 单价列保持数值格式，其余列全部强制文本格式
    text_columns = {idx for idx in range(1, len(headers) + 1) if idx != 5}
    return _export_to_xlsx(
        sheets=[
            ExportSheet(
                title="耗材订单",
                headers=headers,
                rows=rows,
                text_columns=text_columns,
            )
        ],
        filename_prefix="consumable_orders_export",
    )
