# Excel 批量导入与模板生成。
from dataclasses import dataclass
import math
from numbers import Real
import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Tuple, Optional
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from fastapi import HTTPException, status, UploadFile
from app.models.inventory import Inventory, ManualInventoryCreate
from app.services.inventory_status import derive_inventory_quantity_status
from app.services.cas_utils import normalize_cas, validate_cas_format
from app.services.internal_code import (
    INTERNAL_CODE_CONFLICT_MAX_RETRIES,
    build_internal_code_prefix,
    format_internal_code,
    get_max_sequence_for_prefix,
    is_internal_code_unique_violation,
)
from app.services.inventory_operation_logger import (
    SOURCE_BATCH_IMPORT,
    log_stock_in,
)
from app.services.spec_utils import format_specification, parse_specification
from app.services.shelf_utils import normalize_storage_location
from app.services.pinyin_utils import compute_pinyin_fields
from app.core.constants import (
    EXCEL_DATE_EPOCH,
    EXCEL_FILE_MAX_BYTES,
    EXCEL_RED_FONT_COLOR,
    INTERNAL_CODE_MAX_SEQUENCE,
)
from app.core.time_utils import get_utc_now, normalize_to_utc_naive


def _compute_remaining_percent(remaining: Optional[float], initial: Optional[float]) -> Optional[float]:
    if initial is None or initial <= 0:
        return None
    if remaining is None:
        return None
    return remaining / initial


# ==================== 文件上传安全 ====================
# 允许的文件扩展名
ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".csv"}

# 允许的 MIME 类型
ALLOWED_MIME_TYPES = {
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls": "application/vnd.ms-excel",
    ".csv": "text/csv",
}

# 文件魔数（文件头签名）
FILE_MAGIC_BYTES = {
    ".xlsx": b"PK\x03\x04",  # ZIP-based (Office Open XML)
    ".xls": b"\xd0\xcf\x11\xe0",  # OLE2 compound document
    ".csv": b"",  # CSV is text, no magic bytes needed
}

# 最大文件大小 (2MB)
MAX_FILE_SIZE = EXCEL_FILE_MAX_BYTES
BOOLEAN_FALSE_STRINGS = {"false", "0", "no", "n"}
BOOLEAN_TRUE_STRINGS = {"true", "1", "yes", "y"}
EXCEL_IMPORT_COLUMN_MAPPING = {
    "cas_number": ["cas_number", "cas", "cas号"],
    "name": ["name", "名称", "品名"],
    "english_name": ["english_name", "英文名", "englishname"],
    "alias": ["alias", "别名"],
    "category": ["category", "分类", "类别"],
    "brand": ["brand", "品牌", "厂商", "manufacturer"],
    "specification": ["specification", "规格", "spec"],
    "remaining_quantity": ["remaining_quantity", "剩余数量", "剩余量"],
    "storage_location": ["storage_location", "location", "位置", "存放位置"],
    "is_hazardous": ["is_hazardous", "危险品", "是否危险品"],
    "notes": ["notes", "备注", "remark"],
    "created_at": ["created_at", "入库时间", "创建时间", "stock_in_date"],
}


@dataclass
class ExcelImportContext:
    db: Session
    sequence_tracker: dict[tuple[str, str], int]
    default_storage_location: Optional[str]
    default_is_hazardous: bool
    user_id: int


@dataclass
class PreparedInventoryImport:
    total_rows: int
    created_items: list[Inventory]
    errors: list[dict[str, Any]]
    preview_items: list[dict[str, Any]]

    @property
    def valid_rows(self) -> int:
        return len(self.created_items)


REQUIRED_IMPORT_COLUMNS = {"cas_number", "name", "specification"}
REQUIRED_COLUMN_MARKERS = ("（必填）", "(必填)")


def _get_model_field_max_length(model: type, field_name: str) -> int:
    for metadata in model.model_fields[field_name].metadata:
        max_length = getattr(metadata, "max_length", None)
        if isinstance(max_length, int):
            return max_length
    raise RuntimeError(f"Missing max_length metadata for {model.__name__}.{field_name}")


IMPORT_TEXT_MAX_LENGTHS = {
    "name": _get_model_field_max_length(Inventory, "name"),
    "english_name": _get_model_field_max_length(Inventory, "english_name"),
    "alias": _get_model_field_max_length(Inventory, "alias"),
    "category": _get_model_field_max_length(Inventory, "category"),
    "brand": _get_model_field_max_length(Inventory, "brand"),
    "specification": _get_model_field_max_length(ManualInventoryCreate, "specification"),
    "storage_location": _get_model_field_max_length(Inventory, "storage_location"),
    "notes": _get_model_field_max_length(Inventory, "notes"),
}


def validate_uploaded_file(file: UploadFile) -> None:
    # 导入入口先拦截伪造扩展名和超大文件，避免后续解析库处理恶意输入。
    ext = Path(file.filename).suffix.lower() if file.filename else ""
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file type. Only .xlsx, .xls, .csv are allowed"
        )

    file.file.seek(0, 2)  # Seek to end
    file_size = file.file.tell()
    file.file.seek(0)  # Reset to start

    if file_size > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File size exceeds 2MB limit"
        )

    if file_size == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File is empty"
        )

    header = file.file.read(8)
    file.file.seek(0)  # Reset to start

    if ext == ".xlsx":
        if not header.startswith(b"PK\x03\x04"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid XLSX file format"
            )
    elif ext == ".xls":
        if not header.startswith(b"\xd0\xcf\x11\xe0"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid XLS file format"
            )
    # CSV 是纯文本格式，不依赖文件头魔数。


def _parse_boolean(value, default: bool = False) -> bool:
    if pd.api.types.is_bool(value):
        return bool(value)
    if value is None or pd.isna(value):
        return default
    if isinstance(value, Real):
        numeric_value = float(value)
        if numeric_value == 0:
            return False
        if numeric_value == 1:
            return True
        raise ValueError("Invalid is_hazardous: expected true/false or 1/0")
    if not isinstance(value, str):
        raise ValueError("Invalid is_hazardous: expected true/false or 1/0")

    stripped = value.strip().lower()
    if not stripped:
        return default
    if stripped in BOOLEAN_FALSE_STRINGS:
        return False
    if stripped in BOOLEAN_TRUE_STRINGS:
        return True
    raise ValueError("Invalid is_hazardous: expected true/false or 1/0")


class ExcelImportError(Exception):
    def __init__(self, row: int, message: str):
        self.row = row
        self.message = message
        super().__init__(f"Row {row}: {message}")


def _generate_internal_code_with_tracking(
    db: Session,
    cas_number: str,
    sequence_tracker: dict[tuple[str, str], int],
    created_at: Optional[datetime] = None
) -> str:
    # 同一批次里相同 CAS 也要拿到不同流水号，防止内部编码唯一约束冲突。
    date_str = (created_at or get_utc_now()).strftime("%y%m%d")
    prefix = build_internal_code_prefix(cas_number, created_at=created_at)
    tracker_key = (prefix, date_str)

    if tracker_key in sequence_tracker:
        seq = sequence_tracker[tracker_key]
        if seq > INTERNAL_CODE_MAX_SEQUENCE:
            raise ValueError(
                f"Internal code sequence limit reached for {cas_number} on {date_str}: "
                f"max is {INTERNAL_CODE_MAX_SEQUENCE}"
            )
        sequence_tracker[tracker_key] = seq + 1
    else:
        max_seq = get_max_sequence_for_prefix(db, prefix)
        seq = max_seq + 1
        if seq > INTERNAL_CODE_MAX_SEQUENCE:
            raise ValueError(
                f"Internal code sequence limit reached for {cas_number} on {date_str}: "
                f"max is {INTERNAL_CODE_MAX_SEQUENCE}"
            )
        sequence_tracker[tracker_key] = seq + 1

    return format_internal_code(prefix, seq)


def parse_excel_file(file_path: str) -> pd.DataFrame:
    if file_path.endswith(".csv"):
        for encoding in ["utf-8-sig", "utf-8", "gbk", "gb2312"]:
            try:
                return pd.read_csv(file_path, encoding=encoding, keep_default_na=False)
            except UnicodeDecodeError:
                continue
        return pd.read_csv(
            file_path,
            encoding="utf-8-sig",
            encoding_errors="replace",
            keep_default_na=False,
        )
    return pd.read_excel(file_path, keep_default_na=False)


def _normalize_import_text(value: object) -> Optional[str]:
    if value is None or pd.isna(value):
        return None
    normalized = str(value).strip()
    return normalized or None


def _validate_import_text_lengths(row: dict) -> Optional[str]:
    for field, max_length in IMPORT_TEXT_MAX_LENGTHS.items():
        value = _normalize_import_text(row.get(field))
        if field == "storage_location":
            value = normalize_storage_location(value)
        if value is not None and len(value) > max_length:
            return f"Invalid {field}: must not exceed {max_length} characters"
    return None


def _validate_remaining_quantity(row: dict, initial_quantity: float) -> Optional[str]:
    remaining_text = _normalize_import_text(row.get("remaining_quantity"))
    if remaining_text is None:
        return None
    try:
        remaining_value = float(remaining_text)
    except (ValueError, TypeError):
        return "Invalid remaining_quantity: must be a number"
    if not math.isfinite(remaining_value):
        return "Invalid remaining_quantity: must be a finite number"
    if remaining_value < 0:
        return "Invalid remaining_quantity: cannot be negative"
    if remaining_value > initial_quantity:
        return (
            f"Invalid remaining_quantity: {remaining_value} "
            f"cannot exceed initial_quantity {initial_quantity}"
        )
    return None


def validate_row_data(row: dict) -> Tuple[bool, Optional[str]]:
    length_error = _validate_import_text_lengths(row)
    if length_error:
        return False, length_error

    for field in REQUIRED_IMPORT_COLUMNS:
        if _normalize_import_text(row.get(field)) is None:
            return False, f"Missing required field: {field}"

    cas_raw = str(row["cas_number"]).strip()
    normalized_cas = normalize_cas(cas_raw)
    is_valid, error = validate_cas_format(normalized_cas)

    if not is_valid:
        return False, f"Invalid CAS format: {error}"

    try:
        spec_value, _ = parse_specification(str(row["specification"]))
    except ValueError as e:
        return False, f"Invalid specification format: {str(e)}"

    remaining_error = _validate_remaining_quantity(row, spec_value)
    if remaining_error:
        return False, remaining_error

    return True, None


def _parse_import_dataframe(file_path: str) -> pd.DataFrame:
    try:
        return parse_excel_file(file_path)
    except Exception as exc:
        raise ExcelImportError(1, "Failed to parse Excel file") from exc


def _normalize_import_columns(df: pd.DataFrame) -> pd.DataFrame:
    normalized_df = pd.DataFrame()
    for standard_col, possible_cols in EXCEL_IMPORT_COLUMN_MAPPING.items():
        possible_names = {_normalize_import_column_name(candidate) for candidate in possible_cols}
        for col in df.columns:
            if _normalize_import_column_name(col) in possible_names:
                normalized_df[standard_col] = df[col]
                break
    return normalized_df


def _normalize_import_column_name(value: object) -> str:
    normalized = str(value).strip().lower()
    for marker in REQUIRED_COLUMN_MARKERS:
        if normalized.endswith(marker):
            return normalized[: -len(marker)].strip()
    return normalized


def _validate_required_import_columns(df: pd.DataFrame, normalized_df: pd.DataFrame) -> None:
    if df.empty:
        return

    missing_columns = sorted(REQUIRED_IMPORT_COLUMNS - set(normalized_df.columns))
    if missing_columns:
        raise ExcelImportError(1, f"Missing required columns: {', '.join(missing_columns)}")


def _parse_remaining_quantity(row: dict, initial_quantity: float) -> float:
    remaining_qty = initial_quantity
    remaining_raw = row.get("remaining_quantity")
    if pd.notna(remaining_raw):
        remaining_text = str(remaining_raw).strip()
        if remaining_text:
            remaining_qty = float(remaining_text)
    return remaining_qty


def _normalize_import_optional_fields(row: dict, default_storage_location: Optional[str]) -> dict[str, Optional[str]]:
    all_optional_fields = {
        "storage_location": row.get("storage_location"),
        "alias": row.get("alias"),
        "english_name": row.get("english_name"),
        "category": row.get("category"),
        "brand": row.get("brand"),
        "notes": row.get("notes"),
    }
    normalized_optional = {
        field: _normalize_import_text(value)
        for field, value in all_optional_fields.items()
    }
    normalized_optional["storage_location"] = normalize_storage_location(
        normalized_optional["storage_location"] or default_storage_location
    )
    return normalized_optional


def _parse_import_created_at(value: object) -> Optional[datetime]:
    date_str = _normalize_import_text(value)
    if date_str is None:
        return None

    try:
        if date_str.isdigit():
            if len(date_str) == 5:
                date_str = str(EXCEL_DATE_EPOCH + timedelta(days=int(date_str)))
            elif len(date_str) == 8:
                date_str = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
            elif len(date_str) == 6:
                date_str = f"20{date_str[:2]}-{date_str[2:4]}-{date_str[4:6]}"
        parsed_date = pd.to_datetime(date_str)
        if pd.isna(parsed_date):
            raise ValueError("date is missing")
        return parsed_date.to_pydatetime()
    except (ValueError, TypeError, OverflowError) as exc:
        raise ValueError("Invalid created_at: expected a valid date") from exc


def _build_inventory_from_import_row(
    context: ExcelImportContext,
    row: dict,
) -> Inventory:
    normalized_cas = normalize_cas(str(row["cas_number"]))
    spec_value, unit = parse_specification(str(row["specification"]))
    initial_quantity = spec_value
    remaining_qty = _parse_remaining_quantity(row, initial_quantity)
    optional_fields = _normalize_import_optional_fields(row, context.default_storage_location)
    imported_created_at = _parse_import_created_at(row.get("created_at"))
    stored_created_at = normalize_to_utc_naive(imported_created_at)
    internal_code = _generate_internal_code_with_tracking(
        context.db,
        normalized_cas,
        context.sequence_tracker,
        imported_created_at,
    )
    name = str(row["name"]).strip()
    pinyin_fields = compute_pinyin_fields(
        name=name,
        category=optional_fields["category"],
        brand=optional_fields["brand"],
        storage_location=optional_fields["storage_location"],
    )

    return Inventory(
        internal_code=internal_code,
        cas_number=normalized_cas,
        name=name,
        english_name=optional_fields["english_name"],
        alias=optional_fields["alias"],
        category=optional_fields["category"],
        brand=optional_fields["brand"],
        storage_location=optional_fields["storage_location"],
        initial_quantity=initial_quantity,
        remaining_quantity=remaining_qty,
        remaining_percent=_compute_remaining_percent(remaining_qty, initial_quantity),
        unit=unit,
        is_hazardous=_parse_boolean(row.get("is_hazardous"), context.default_is_hazardous),
        status=derive_inventory_quantity_status(remaining_qty, initial_quantity),
        notes=optional_fields["notes"],
        created_at=stored_created_at,
        created_by_id=context.user_id,
        **pinyin_fields,
    )


def _serialize_import_preview_item(row_num: int, inventory: Inventory) -> dict[str, Any]:
    return {
        "row": row_num,
        "cas_number": inventory.cas_number,
        "name": inventory.name,
        "brand": inventory.brand,
        "category": inventory.category,
        "specification": format_specification(inventory.initial_quantity, inventory.unit),
        "remaining_quantity": inventory.remaining_quantity,
        "storage_location": inventory.storage_location,
    }


def _build_import_result(
    prepared: PreparedInventoryImport,
    *,
    created: int,
) -> dict[str, Any]:
    cas_numbers = sorted({item.cas_number for item in prepared.created_items if item.cas_number})
    return {
        "success": len(prepared.errors) == 0,
        "total_rows": prepared.total_rows,
        "valid_rows": prepared.valid_rows,
        "created": created,
        "created_cas_numbers": cas_numbers if created > 0 else [],
        "errors": prepared.errors,
        "preview_items": prepared.preview_items,
    }


def _prepare_inventory_import(
    db: Session,
    file_path: str,
    default_storage_location: Optional[str] = None,
    default_is_hazardous: bool = False,
    user_id: int = 1,
) -> PreparedInventoryImport:
    df = _parse_import_dataframe(file_path)
    normalized_df = _normalize_import_columns(df)
    _validate_required_import_columns(df, normalized_df)
    errors: list[dict[str, Any]] = []
    preview_items: list[dict[str, Any]] = []
    created_items: list[Inventory] = []
    sequence_tracker: dict[tuple[str, str], int] = {}
    import_context = ExcelImportContext(
        db=db,
        sequence_tracker=sequence_tracker,
        default_storage_location=default_storage_location,
        default_is_hazardous=default_is_hazardous,
        user_id=user_id,
    )

    for idx, row in normalized_df.iterrows():
        row_num = idx + 2
        is_valid, error = validate_row_data(row)
        if not is_valid:
            errors.append({"row": row_num, "error": error})
            continue
        try:
            inventory = _build_inventory_from_import_row(
                import_context,
                row,
            )
        except Exception as exc:
            errors.append({"row": row_num, "error": str(exc)})
            continue

        created_items.append(inventory)
        preview_items.append(_serialize_import_preview_item(row_num, inventory))

    return PreparedInventoryImport(
        total_rows=len(normalized_df),
        created_items=created_items,
        errors=errors,
        preview_items=preview_items,
    )


def _persist_imported_inventory(
    db: Session,
    *,
    created_items: list[Inventory],
    user_id: int,
    is_cli: bool,
) -> None:
    with db.begin_nested():
        for inventory in created_items:
            db.add(inventory)

        db.flush()
        for inventory in created_items:
            log_stock_in(
                db,
                inventory=inventory,
                operator_id=user_id,
                source=SOURCE_BATCH_IMPORT,
                is_cli=is_cli,
            )
        db.flush()

    db.commit()


def preview_inventory_import_from_excel(
    db: Session,
    file_path: str,
    default_storage_location: Optional[str] = None,
    default_is_hazardous: bool = False,
    user_id: int = 1,
) -> dict[str, Any]:
    prepared = _prepare_inventory_import(
        db=db,
        file_path=file_path,
        default_storage_location=default_storage_location,
        default_is_hazardous=default_is_hazardous,
        user_id=user_id,
    )
    db.rollback()
    return _build_import_result(prepared, created=0)


def confirm_inventory_import_from_excel(
    db: Session,
    file_path: str,
    default_storage_location: Optional[str] = None,
    default_is_hazardous: bool = False,
    user_id: int = 1,
    is_cli: bool = False,
) -> dict[str, Any]:
    # 批量导入保持全有或全无；确认阶段会重新完整校验，避免预览后文件变化导致脏写入。
    for attempt in range(INTERNAL_CODE_CONFLICT_MAX_RETRIES):
        prepared = _prepare_inventory_import(
            db=db,
            file_path=file_path,
            default_storage_location=default_storage_location,
            default_is_hazardous=default_is_hazardous,
            user_id=user_id,
        )
        if prepared.errors:
            db.rollback()
            return _build_import_result(prepared, created=0)

        try:
            _persist_imported_inventory(
                db,
                created_items=prepared.created_items,
                user_id=user_id,
                is_cli=is_cli,
            )
            return _build_import_result(prepared, created=prepared.valid_rows)
        except IntegrityError as exc:
            if not is_internal_code_unique_violation(exc):
                db.rollback()
                raise Exception(f"Failed to save imported data: {str(exc)}") from exc
            db.rollback()
            if attempt == INTERNAL_CODE_CONFLICT_MAX_RETRIES - 1:
                raise Exception("Failed to allocate unique internal codes during import, please retry") from exc
        except Exception as exc:
            db.rollback()
            raise Exception(f"Failed to save imported data: {str(exc)}") from exc

    raise Exception("Failed to import inventory after retries")


def generate_excel_template() -> bytes:
    # 所有列强制文本格式，避免 Excel 把 CAS 号自动改成日期。
    from io import BytesIO

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.styles.numbers import FORMAT_TEXT
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        from fastapi import HTTPException, status
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel template generation requires the 'openpyxl' package. "
                   "Please ensure openpyxl is installed in the production environment.",
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "库存导入模板"

    # 定义列: (列键, 中文标签, 示例值)
    columns = [
        ("cas_number", "CAS号（必填）", "64-17-5"),
        ("name", "名称（必填）", "乙醇"),
        ("english_name", "英文名", "Ethanol"),
        ("alias", "别名", "酒精"),
        ("category", "分类", "有机溶剂"),
        ("brand", "品牌", "Sigma"),
        ("specification", "规格（必填）", "500mL"),
        ("remaining_quantity", "剩余量", ""),
        ("storage_location", "存放位置", "2-6-6-1"),
        ("is_hazardous", "是否危险品", ""),
        ("notes", "备注", "请删除示例数据"),
    ]

    # 定义表头字体样式
    header_font = Font(bold=True)

    # 示例数据使用红色字体
    example_font = Font(color=EXCEL_RED_FONT_COLOR)  # 红色

    # 写入表头和示例数据，同时设置列格式和列宽（单次遍历优化性能）
    for col_idx, (key, label, example) in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)

        # 设置表头
        header_cell = ws.cell(row=1, column=col_idx, value=label)
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cell.number_format = FORMAT_TEXT

        # 写入示例数据（红色字体）
        example_cell = ws.cell(row=2, column=col_idx, value=example)
        example_cell.font = example_font
        example_cell.number_format = FORMAT_TEXT

        # 列样式用于后续新建单元格，已存在的单元格需在上方显式设置
        ws.column_dimensions[col_letter].number_format = FORMAT_TEXT
        ws.column_dimensions[col_letter].width = 15 if key == "cas_number" else 12

    # 保存到字节流
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_import_template() -> dict:
    return {
        "columns": [
            {
                "name": "cas_number",
                "label": "CAS号",
                "required": True,
                "description": "格式: XXXXX-XX-X，例如 64-17-5"
            },
            {
                "name": "name",
                "label": "名称",
                "required": True,
                "description": "化学品中文名称，例如 乙醇"
            },
            {
                "name": "english_name",
                "label": "英文名",
                "required": False,
                "description": "化学品的英文名称，例如 Ethanol"
            },
            {
                "name": "alias",
                "label": "别名",
                "required": False,
                "description": "化学品的别名或俗称，例如 酒精"
            },
            {
                "name": "category",
                "label": "分类",
                "required": False,
                "description": "化学品分类，例如 有机溶剂、酸、碱"
            },
            {
                "name": "brand",
                "label": "品牌/厂商",
                "required": False,
                "description": "品牌或生产厂家，例如 Sigma、阿拉丁"
            },
            {
                "name": "specification",
                "label": "规格",
                "required": True,
                "description": "格式: 数值+单位，如 500mL, 1L, 100g，系统会自动解析出数量和单位"
            },
            {
                "name": "remaining_quantity",
                "label": "剩余量",
                "required": False,
                "description": "剩余量（可选），不填则默认等于规格中的数量"
            },
            {
                "name": "storage_location",
                "label": "存放位置",
                "required": False,
                "description": "例如 302冰箱第二层、A-1-1 柜"
            },
            {
                "name": "is_hazardous",
                "label": "是否危险品",
                "required": False,
                "description": "true/false 或 1/0，危险品需要特殊存储"
            },
            {
                "name": "notes",
                "label": "备注",
                "required": False,
                "description": "其他需要记录的信息，例如 易燃物品"
            },
            {
                "name": "created_at",
                "label": "入库日期",
                "required": False,
                "description": "支持格式: YYYY-MM-DD、YYYY/MM/DD、YYMMDD、YYYYMMDD、Excel序列号(如 45292)。留空则使用导入时间"
            }
        ]
    }
